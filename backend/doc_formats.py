"""
doc_formats.py — universal document text extraction for heterogeneous formats.

Supported:
  .txt, .md          plain text
  .pdf               PyPDF2 / pypdf text layer (+ OCR fallback for scanned pages)
  .xlsx, .xls        openpyxl spreadsheet rows
  .csv               tabular text (also ingested as sensor data separately)
  .eml, .msg-like    email archives (stdlib email parser)
  .png, .jpg, .jpeg  OCR via pytesseract (optional)
  .pid, .p&id        P&ID tag-list / drawing exports (text + tag extraction)

Returns normalized plain text suitable for the existing parser metadata extractors.
"""
from __future__ import annotations

import csv
import email
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

# Equipment tag pattern common on P&IDs and industrial drawings
_EQUIP_TAG_RE = re.compile(
    r"\b([A-Z]{2,5}[-/]?\d{1,4}[A-Z]?)\b"
)

_TEXT_EXTENSIONS = {".txt", ".md", ".log", ".pid", ".p&id"}
_PDF_EXTENSIONS = {".pdf"}
_SHEET_EXTENSIONS = {".xlsx", ".xls"}
_CSV_EXTENSIONS = {".csv"}
_EMAIL_EXTENSIONS = {".eml"}
_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}


@dataclass
class ExtractedDocument:
    source_path: str
    format_type: str          # txt | pdf | spreadsheet | email | ocr | pid | csv_doc
    raw_text: str
    metadata: dict            # extra fields (sheet names, email headers, pid tags, …)


def _read_text(path: Path) -> str:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(path: Path) -> tuple[str, dict]:
    text_parts: list[str] = []
    meta: dict = {"pages": 0, "ocr_pages": []}

    # Primary: text layer via pypdf
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        meta["pages"] = len(reader.pages)
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text() or ""
            if page_text.strip():
                text_parts.append(f"--- PAGE {i + 1} ---\n{page_text}")
            else:
                meta["ocr_pages"].append(i + 1)
    except ImportError:
        try:
            import PyPDF2
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                meta["pages"] = len(reader.pages)
                for i, page in enumerate(reader.pages):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        text_parts.append(f"--- PAGE {i + 1} ---\n{page_text}")
                    else:
                        meta["ocr_pages"].append(i + 1)
        except ImportError:
            return "", {"error": "pypdf not installed"}

    # OCR fallback for scanned pages
    if meta.get("ocr_pages"):
        ocr_text = _ocr_pdf_pages(path, meta["ocr_pages"])
        if ocr_text:
            text_parts.append("--- OCR EXTRACTED ---\n" + ocr_text)
            meta["ocr_used"] = True

    return "\n\n".join(text_parts), meta


def _ocr_pdf_pages(path: Path, page_numbers: list[int]) -> str:
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except ImportError:
        return ""

    try:
        images = convert_from_path(str(path), first_page=min(page_numbers),
                                   last_page=max(page_numbers))
    except Exception:
        return ""

    chunks = []
    for img in images:
        try:
            chunks.append(pytesseract.image_to_string(img))
        except Exception:
            continue
    return "\n".join(chunks)


def _extract_image_ocr(path: Path) -> tuple[str, dict]:
    try:
        from PIL import Image
        import pytesseract
        img = Image.open(path)
        text = pytesseract.image_to_string(img)
        tags = sorted(set(_EQUIP_TAG_RE.findall(text.upper())))
        return text, {"ocr": True, "pid_tags": tags}
    except ImportError:
        return f"[OCR unavailable — install pillow + pytesseract for {path.name}]", {"ocr": False}
    except Exception as e:
        return f"[OCR failed: {e}]", {"ocr": False, "error": str(e)}


def _extract_spreadsheet(path: Path) -> tuple[str, dict]:
    suffix = path.suffix.lower()
    meta: dict = {"sheets": []}

    if suffix == ".csv":
        rows = []
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append("\t".join(row))
        return "\n".join(rows), {"format": "csv"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            meta["sheets"].append(sheet_name)
            ws = wb[sheet_name]
            parts.append(f"--- SHEET: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts), meta
    except ImportError:
        return f"[openpyxl not installed — cannot read {path.name}]", {"error": "openpyxl missing"}


def _extract_email(path: Path) -> tuple[str, dict]:
    raw = path.read_bytes()
    msg = email.message_from_bytes(raw)
    headers = {
        "subject": msg.get("Subject", ""),
        "from":    msg.get("From", ""),
        "to":      msg.get("To", ""),
        "date":    msg.get("Date", ""),
    }
    body_parts = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    body_parts.append(payload.decode("utf-8", errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            body_parts.append(payload.decode("utf-8", errors="replace"))

    header_block = "\n".join(f"{k.upper()}: {v}" for k, v in headers.items() if v)
    text = header_block + "\n\n" + "\n".join(body_parts)
    return text, {"email": headers}


def _extract_pid(path: Path, base_text: str) -> tuple[str, dict]:
    """P&ID exports: enrich text with extracted equipment tag inventory."""
    tags = sorted(set(_EQUIP_TAG_RE.findall(base_text.upper())))
    tag_block = ""
    if tags:
        tag_block = "\n\n--- P&ID EQUIPMENT TAG INVENTORY ---\n" + ", ".join(tags)
    return base_text + tag_block, {"pid_tags": tags, "format": "pid"}


def extract_text(path: str | Path) -> ExtractedDocument:
    path = Path(path)
    suffix = path.suffix.lower()
    stem_lower = path.stem.lower()

    if suffix in _TEXT_EXTENSIONS or suffix == "":
        raw = _read_text(path)
        fmt = "pid" if ("pid" in stem_lower or "p&id" in stem_lower) else "txt"
        if fmt == "pid":
            raw, meta = _extract_pid(path, raw)
        else:
            meta = {}
        return ExtractedDocument(str(path.resolve()), fmt, raw, meta)

    if suffix in _PDF_EXTENSIONS:
        raw, meta = _extract_pdf(path)
        tags = sorted(set(_EQUIP_TAG_RE.findall(raw.upper())))
        if tags:
            meta["pid_tags"] = tags
            raw += "\n\n--- EXTRACTED EQUIPMENT TAGS ---\n" + ", ".join(tags)
        return ExtractedDocument(str(path.resolve()), "pdf", raw, meta)

    if suffix in _SHEET_EXTENSIONS:
        raw, meta = _extract_spreadsheet(path)
        return ExtractedDocument(str(path.resolve()), "spreadsheet", raw, meta)

    if suffix in _CSV_EXTENSIONS:
        # Non-sensor CSVs treated as tabular documents
        raw, meta = _extract_spreadsheet(path)
        return ExtractedDocument(str(path.resolve()), "csv_doc", raw, meta)

    if suffix in _EMAIL_EXTENSIONS:
        raw, meta = _extract_email(path)
        return ExtractedDocument(str(path.resolve()), "email", raw, meta)

    if suffix in _IMAGE_EXTENSIONS:
        raw, meta = _extract_image_ocr(path)
        if "pid" in stem_lower or "drawing" in stem_lower:
            raw, meta = _extract_pid(path, raw)
            fmt = "pid"
        else:
            fmt = "ocr"
        return ExtractedDocument(str(path.resolve()), fmt, raw, meta)

    # Unknown — try as text
    try:
        raw = _read_text(path)
        return ExtractedDocument(str(path.resolve()), "unknown", raw, {})
    except Exception as e:
        return ExtractedDocument(str(path.resolve()), "unsupported", "", {"error": str(e)})


def supported_extensions() -> set[str]:
    return (
        _TEXT_EXTENSIONS | _PDF_EXTENSIONS | _SHEET_EXTENSIONS
        | _CSV_EXTENSIONS | _EMAIL_EXTENSIONS | _IMAGE_EXTENSIONS
    )


def scan_directory(docs_dir: str | Path) -> list[Path]:
    docs_dir = Path(docs_dir)
    exts = supported_extensions()
    files: list[Path] = []
    for p in sorted(docs_dir.rglob("*")):
        if p.is_file() and p.suffix.lower() in exts:
            # Skip sensor CSVs — handled by fact_builder.load_sensor_csvs
            if p.suffix.lower() == ".csv" and "sensor" in p.parts:
                continue
            files.append(p)
    return files
